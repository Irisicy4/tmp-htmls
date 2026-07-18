#!/usr/bin/env python3
"""Generate per-task review packets for curator review.

For each task with a tests/test_grounded.py:
  - <task_dir>/review_full_prompt.txt — full agent prompt (instruction.md
    + the schema-rendered inline addendum)
  - <task_dir>/review.xlsx — structured worksheet with four review
    sections and a task-level verdict footer:
      Section 1 — Instruction text
      Section 2 — JSON Fields (one row per required/optional field)
      Section 3 — Hard Constraints (one row per predicate)
      Section 4 — Web Grounding Checks (one row per URL+claim pair)
      Footer    — task-level Passed? / Abandon?

Usage:
    python scripts_batch/build_review_packets.py [TASK_DIR_OR_GLOB]
    python scripts_batch/build_review_packets.py tasks/real_118/task-01-*
    python scripts_batch/build_review_packets.py --all
"""
from __future__ import annotations

import argparse
import importlib.util
import inspect
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
CHECKBOX_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FONT = Font(bold=True, size=12, color="1F3864")
TASK_HEADER_FONT = Font(bold=True, size=14, color="1F3864")


def _load_spec(spec_path: Path):
    mod_name = f"_review_{spec_path.parent.parent.name.replace('-', '_')}"
    spec = importlib.util.spec_from_file_location(mod_name, spec_path)
    mod = importlib.util.module_from_spec(spec)
    sys.modules[mod_name] = mod
    mod.__name__ = mod_name
    spec.loader.exec_module(mod)
    return mod


def _describe_constraint(c) -> tuple[str, str]:
    """Return (rule_name, human-readable summary) for a constraint object."""
    name = getattr(c, "name", "") or type(c).__name__
    pairs = []
    for k, v in c.__dict__.items():
        if k == "name":
            continue
        v_repr = repr(v)
        if len(v_repr) > 60:
            v_repr = v_repr[:57] + "…"
        pairs.append(f"{k}={v_repr}")
    return name, f"{type(c).__name__}({', '.join(pairs)})"


def _hard_constraint_paths(c) -> set[str]:
    """Best-effort: which schema paths does this constraint touch?"""
    paths = set()
    for k in ("path", "list_path", "required", "sub_path"):
        v = getattr(c, k, None)
        if isinstance(v, str):
            if k == "sub_path":
                lp = getattr(c, "list_path", "")
                if lp:
                    paths.add(f"{lp}[].{v}")
            else:
                paths.add(v)
        elif isinstance(v, (list, tuple)):
            for s in v:
                if isinstance(s, str):
                    paths.add(s)
    return paths


def _faithfulness_paths(spec_module) -> set[str]:
    """Source-grep FAITHFULNESS_CHECKS for schema-path mentions.
    We don't try to evaluate; we just look at the function source for
    string keys/paths it cites."""
    paths = set()
    func = getattr(spec_module, "FAITHFULNESS_CHECKS", None)
    if func is None:
        return paths
    try:
        src = inspect.getsource(func)
    except (OSError, TypeError):
        return paths
    import re
    # Catch summary["foo"]["bar"][i]["baz"] → "foo[].bar[].baz" patterns
    for m in re.finditer(r"summary\s*((?:\[(?:[^\]]+)\])+)", src):
        chain = m.group(1)
        parts = re.findall(r"\[([^\]]+)\]", chain)
        out = []
        for p in parts:
            p = p.strip()
            if p.startswith(("'", '"')):
                out.append(p.strip("'\""))
            else:
                # likely a loop index — represents list iteration
                if out:
                    out[-1] = out[-1] + "[]"
        if out:
            paths.add(".".join(out))
    return paths


def _try_faith_checks(spec_module) -> list[dict]:
    """Call FAITHFULNESS_CHECKS on the schema examples to extract concrete
    URL+claim pairs for display. Returns [] on any error."""
    try:
        return list(spec_module.FAITHFULNESS_CHECKS(spec_module.SUMMARY_SCHEMA.examples))[:30]
    except Exception:
        return []


def _row_height(text: str, base: int = 18) -> int:
    if not text:
        return base
    lines = text.count("\n") + 1
    return min(base * max(lines, 1), 200)


def _set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_packet(task_dir: Path, force: bool = False) -> bool:
    """Build review_full_prompt.txt and review.xlsx for one task.
    Returns True if anything was written, False on skip/error."""
    spec_path = task_dir / "tests" / "test_grounded.py"
    if not spec_path.is_file():
        return False
    instruction_path = task_dir / "instruction.md"
    if not instruction_path.is_file():
        return False

    spec_module = _load_spec(spec_path)
    schema = getattr(spec_module, "SUMMARY_SCHEMA", None)
    hards = list(getattr(spec_module, "HARD_CONSTRAINTS", []) or [])

    # ----- review_full_prompt.txt --------------------------------------
    instruction = instruction_path.read_text()
    brief = instruction.split("\n---\n", 1)[0].rstrip() + "\n"
    addendum = schema.to_instruction_addendum(mode="inline") if schema else ""
    full_prompt = brief + addendum
    (task_dir / "review_full_prompt.txt").write_text(full_prompt)

    # ----- review.xlsx --------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Review"
    _set_col_widths(ws, {"A": 36, "B": 22, "C": 14, "D": 16, "E": 16, "F": 12})

    row = 1
    ws.cell(row=row, column=1, value=f"Task review — {task_dir.name}").font = TASK_HEADER_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1, value="Curator: tick ☑ in the YES columns. Add free-text in the 'regenerate instruction' cell when checked.").font = Font(italic=True, size=9, color="595959")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    YES_NO = DataValidation(type="list", formula1='"☐,☑"', allow_blank=True)
    YES_NO.error = "Use ☐ or ☑"
    ws.add_data_validation(YES_NO)

    def _checkbox(cell):
        cell.value = "☐"
        cell.alignment = Alignment(horizontal="center")
        cell.fill = CHECKBOX_FILL
        YES_NO.add(cell)

    # ===== Section 1 — Instruction text =================================
    s = ws.cell(row=row, column=1, value="① Instruction (instruction.md)")
    s.font = SECTION_FONT
    s.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    headers = ["Refers to", "Manual changed?", "Regenerate?", "Regenerate instruction (text →)", "", ""]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1
    ws.cell(row=row, column=1, value="instruction.md (see review_full_prompt.txt for rendered version)")
    _checkbox(ws.cell(row=row, column=2))
    _checkbox(ws.cell(row=row, column=3))
    ws.cell(row=row, column=4, value="")
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    row += 2

    # ===== Section 2 — JSON Fields ======================================
    s = ws.cell(row=row, column=1, value="② JSON Fields (SUMMARY_SCHEMA)")
    s.font = SECTION_FONT; s.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1, value="Section-level:").font = Font(italic=True)
    ws.cell(row=row, column=2, value="Regenerate all?")
    _checkbox(ws.cell(row=row, column=3))
    ws.cell(row=row, column=4, value="Regenerate instruction →")
    ws.cell(row=row, column=5, value="")
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    row += 1

    headers = ["Path", "Type", "Required?", "Hard-constrained?", "Grounding URL?", "Delete?"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row += 1

    constraint_paths: set[str] = set()
    for c in hards:
        constraint_paths |= _hard_constraint_paths(c)
    faith_paths = _faithfulness_paths(spec_module)

    if schema:
        for f in schema.required:
            ws.cell(row=row, column=1, value=f.path)
            ws.cell(row=row, column=2, value=f.type_hint)
            ws.cell(row=row, column=3, value="YES").alignment = Alignment(horizontal="center")
            hc = "YES" if any(f.path == p or f.path.startswith(p + ".") or p.startswith(f.path) for p in constraint_paths) else ""
            gu = "YES" if (any(f.path == p or f.path.startswith(p + ".") or p.startswith(f.path) for p in faith_paths) or
                           "url" in f.path.lower() or "source" in f.path.lower()) else ""
            ws.cell(row=row, column=4, value=hc).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=5, value=gu).alignment = Alignment(horizontal="center")
            _checkbox(ws.cell(row=row, column=6))
            row += 1
        for f in schema.optional:
            ws.cell(row=row, column=1, value=f.path)
            ws.cell(row=row, column=2, value=f.type_hint)
            ws.cell(row=row, column=3, value="optional").alignment = Alignment(horizontal="center")
            hc = "YES" if any(f.path == p or f.path.startswith(p + ".") or p.startswith(f.path) for p in constraint_paths) else ""
            gu = "YES" if (any(f.path == p or f.path.startswith(p + ".") or p.startswith(f.path) for p in faith_paths) or
                           "url" in f.path.lower() or "source" in f.path.lower()) else ""
            ws.cell(row=row, column=4, value=hc).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=5, value=gu).alignment = Alignment(horizontal="center")
            _checkbox(ws.cell(row=row, column=6))
            row += 1
    row += 1

    # ===== Section 3 — Hard Constraints =================================
    s = ws.cell(row=row, column=1, value="③ Hard Constraints (HARD_CONSTRAINTS)")
    s.font = SECTION_FONT; s.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1, value="Section-level:").font = Font(italic=True)
    ws.cell(row=row, column=2, value="Regenerate all?")
    _checkbox(ws.cell(row=row, column=3))
    ws.cell(row=row, column=4, value="Regenerate instruction →")
    ws.cell(row=row, column=5, value="")
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    row += 1

    headers = ["Rule name", "Predicate", "", "", "", "Delete?"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        if h:
            c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
    row += 1
    for c in hards:
        name, descr = _describe_constraint(c)
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=descr)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = _row_height(descr, base=18)
        _checkbox(ws.cell(row=row, column=6))
        row += 1
    row += 1

    # ===== Section 4 — Web Grounding Checks =============================
    s = ws.cell(row=row, column=1, value="④ Web Grounding Checks (FAITHFULNESS_CHECKS)")
    s.font = SECTION_FONT; s.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1, value="Section-level:").font = Font(italic=True)
    ws.cell(row=row, column=2, value="Regenerate all?")
    _checkbox(ws.cell(row=row, column=3))
    ws.cell(row=row, column=4, value="Regenerate instruction →")
    ws.cell(row=row, column=5, value="")
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    row += 1

    headers = ["URL (from example summary)", "Claim", "", "", "", "Delete?"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        if h:
            c.font = HEADER_FONT; c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
    row += 1
    faith_samples = _try_faith_checks(spec_module)
    if faith_samples:
        for chk in faith_samples:
            ws.cell(row=row, column=1, value=str(chk.get("url", ""))[:200])
            ws.cell(row=row, column=2, value=str(chk.get("claim", ""))[:300])
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            _checkbox(ws.cell(row=row, column=6))
            row += 1
    else:
        ws.cell(row=row, column=1, value="(FAITHFULNESS_CHECKS is dynamic; no concrete pairs derivable from example summary)").font = Font(italic=True, color="808080")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    row += 1

    # ===== Footer — task-level verdict ==================================
    s = ws.cell(row=row, column=1, value="Task-level verdict")
    s.font = SECTION_FONT; s.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    ws.cell(row=row, column=1, value="Passed?")
    _checkbox(ws.cell(row=row, column=2))
    ws.cell(row=row, column=3, value="Abandon?")
    _checkbox(ws.cell(row=row, column=4))
    ws.cell(row=row, column=5, value="Curator notes →")
    ws.cell(row=row, column=6, value="")
    row += 1

    out = task_dir / "review.xlsx"
    wb.save(out)
    return True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("target", nargs="?", help="Task dir or glob (default: requires --all)")
    ap.add_argument("--all", action="store_true", help="Build for every task in tasks/")
    args = ap.parse_args()

    if args.all:
        roots = ["real_118", "long_horizon/web_research", "long_horizon/stem_research"]
        task_dirs = []
        for root in roots:
            base = REPO_ROOT / "tasks" / root
            task_dirs += sorted(base.glob("task-*"))
    elif args.target:
        p = Path(args.target)
        if p.is_dir():
            task_dirs = [p]
        else:
            from glob import glob
            task_dirs = [Path(x) for x in glob(args.target)]
    else:
        ap.error("provide a task dir, --all, or a glob")

    n_built = 0
    n_skipped = 0
    for td in task_dirs:
        try:
            if build_packet(td):
                n_built += 1
                print(f"✓ {td}")
            else:
                n_skipped += 1
        except Exception as exc:
            n_skipped += 1
            print(f"✗ {td}: {type(exc).__name__}: {exc}")
    print(f"\nbuilt: {n_built}  skipped: {n_skipped}")


if __name__ == "__main__":
    main()
