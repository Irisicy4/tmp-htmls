#!/usr/bin/env python3
"""
Experiment visualization tool.

Generates HTML dashboards and Excel spreadsheets from experiment result directories.

Usage:
    # Single experiment
    python visualize_results.py results/baseline-gpt/

    # Compare N experiments
    python visualize_results.py results/baseline-gpt/ results/skill-gpt/ results/claude/

    # Specify output location
    python visualize_results.py results/exp1/ results/exp2/ --output comparison/
"""

import argparse
import json
import html as html_mod
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional


def load_experiment(exp_dir: Path) -> Dict[str, Any]:
    """Load all result JSONs and experiment metadata from a directory."""
    meta_path = exp_dir / "_experiment.json"
    meta = {}
    if meta_path.exists():
        with open(meta_path) as f:
            meta = json.load(f)

    tasks = {}
    for jf in sorted(exp_dir.glob("*.json")):
        if jf.name.startswith("_"):
            continue
        with open(jf) as f:
            tasks[jf.stem] = json.load(f)

    if not meta.get("experiment_name"):
        meta["experiment_name"] = exp_dir.name

    return {"meta": meta, "tasks": tasks, "dir": str(exp_dir)}


def _format_conversation(conversation: List[Dict[str, Any]]) -> str:
    """Format LLM conversation history into a readable trace string."""
    lines = []
    for msg in conversation:
        role = msg.get("role", "?").upper()
        tool_calls = msg.get("tool_calls", [])
        content = msg.get("content", "") or ""

        if role == "USER":
            # Truncate the initial system prompt (it's huge)
            if isinstance(content, str) and len(content) > 500:
                # Show just the task instruction portion
                if "Task:" in content:
                    task_start = content.index("Task:")
                    task_chunk = content[task_start:task_start + 300]
                    lines.append(f"[USER] {task_chunk}...")
                else:
                    lines.append(f"[USER] {content[:300]}...")
            elif isinstance(content, list):
                # Multimodal content
                text_parts = [p.get("text", "")[:200] for p in content if p.get("type") == "text"]
                img_count = sum(1 for p in content if p.get("type") == "image_url")
                text = " ".join(text_parts)
                suffix = f" [+{img_count} image(s)]" if img_count else ""
                lines.append(f"[USER] {text[:300]}{suffix}")
            else:
                lines.append(f"[USER] {content}")
        elif role == "ASSISTANT":
            if tool_calls:
                for tc in tool_calls:
                    fn = tc.get("function", {})
                    name = fn.get("name", "?")
                    args = fn.get("arguments", "")
                    if len(args) > 200:
                        args = args[:200] + "..."
                    lines.append(f"[ASSISTANT] tool_call: {name}({args})")
            elif content:
                lines.append(f"[ASSISTANT] {content[:500]}")
        elif role == "TOOL":
            tool_id = msg.get("tool_call_id", "")
            if isinstance(content, str) and len(content) > 300:
                lines.append(f"[TOOL {tool_id[:12]}] {content[:300]}...")
            else:
                lines.append(f"[TOOL {tool_id[:12]}] {content}")

    return "\n\n".join(lines)


def extract_task_metrics(result: Dict[str, Any]) -> Dict[str, Any]:
    """Pull key metrics from a single task result JSON."""
    eval_info = result.get("eval", {}) or {}
    details = eval_info.get("details", {}) or {}
    dims = details.get("dimension_scores", {}) or {}
    cost_stats = result.get("api_cost_stats", {}) or {}

    total_cost = 0.0
    if isinstance(cost_stats, dict):
        total_cost = cost_stats.get("total_cost_usd", cost_stats.get("total_cost", 0.0))
    elif isinstance(cost_stats, list):
        total_cost = sum(c.get("cost", 0.0) for c in cost_stats)

    # Token usage
    input_tokens = cost_stats.get("total_input_tokens", 0) if isinstance(cost_stats, dict) else 0
    output_tokens = cost_stats.get("total_output_tokens", 0) if isinstance(cost_stats, dict) else 0
    cached_tokens = cost_stats.get("total_cached_tokens", 0) if isinstance(cost_stats, dict) else 0
    total_tokens = cost_stats.get("total_tokens", input_tokens + output_tokens) if isinstance(cost_stats, dict) else 0
    api_calls = cost_stats.get("api_calls", 0) if isinstance(cost_stats, dict) else 0

    # Build a readable conversation trace from the message history
    conversation = result.get("conversation", [])
    conversation_trace = _format_conversation(conversation)

    return {
        "status": result.get("status", "unknown"),
        "passed": eval_info.get("passed", False),
        "score": details.get("overall_score", None),
        "dimensions": dims,
        "iterations": result.get("iterations", 0),
        "execution_time": result.get("execution_time", 0),
        "cost": total_cost,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "cached_tokens": cached_tokens,
        "total_tokens": total_tokens,
        "api_calls": api_calls,
        "task_result": result.get("task_result", ""),
        "execution_summary": result.get("execution_summary", ""),
        "conversation_trace": conversation_trace,
        "judge_reasoning": details.get("judge_reasoning", ""),
        "skills_used": result.get("skills_used", []),
    }


# ---------------------------------------------------------------------------
# HTML generation
# ---------------------------------------------------------------------------

_CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
       background: #f5f7fa; color: #1a1a2e; padding: 24px; }
h1 { font-size: 1.6rem; margin-bottom: 8px; }
h2 { font-size: 1.2rem; margin: 24px 0 12px; color: #333; }
.header { background: #fff; padding: 20px 24px; border-radius: 12px;
          box-shadow: 0 1px 3px rgba(0,0,0,.08); margin-bottom: 20px; }
.cards { display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 20px; }
.card { background: #fff; border-radius: 10px; padding: 16px 20px;
        box-shadow: 0 1px 3px rgba(0,0,0,.08); min-width: 150px; flex: 1; }
.card .label { font-size: .75rem; color: #888; text-transform: uppercase; letter-spacing: .5px; }
.card .value { font-size: 1.5rem; font-weight: 700; margin-top: 4px; }
table { width: 100%; border-collapse: collapse; background: #fff;
        border-radius: 10px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,.08); }
th { background: #f0f2f5; padding: 10px 12px; text-align: left; font-size: .8rem;
     text-transform: uppercase; letter-spacing: .5px; color: #555; cursor: pointer;
     user-select: none; white-space: nowrap; }
th:hover { background: #e4e7eb; }
td { padding: 10px 12px; border-top: 1px solid #eee; font-size: .85rem; vertical-align: top; }
tr:hover { background: #fafbfc; }
.pass { color: #16a34a; font-weight: 600; }
.fail { color: #dc2626; font-weight: 600; }
.error { color: #ea580c; font-weight: 600; }
.delta-pos { color: #16a34a; }
.delta-neg { color: #dc2626; }
.expandable { cursor: pointer; }
.expand-icon { display: inline-block; width: 16px; transition: transform .2s; }
.expand-icon.open { transform: rotate(90deg); }
.detail-row { display: none; }
.detail-row.open { display: table-row; }
.detail-cell { padding: 16px 24px; background: #fafbfc; }
.detail-cell pre { white-space: pre-wrap; word-break: break-word;
                   background: #f0f2f5; padding: 12px; border-radius: 6px;
                   font-size: .8rem; max-height: 400px; overflow-y: auto; margin-top: 8px; }
.detail-cell h4 { margin: 12px 0 4px; font-size: .85rem; color: #555; }
.detail-cell h4:first-child { margin-top: 0; }
.side-by-side { display: grid; gap: 16px; }
.exp-label { font-size: .75rem; color: #888; font-weight: 600;
             text-transform: uppercase; margin-bottom: 4px; }
.filter-bar { margin-bottom: 12px; display: flex; gap: 8px; align-items: center; }
.filter-bar button { padding: 4px 12px; border: 1px solid #ddd; border-radius: 6px;
                     background: #fff; cursor: pointer; font-size: .8rem; }
.filter-bar button.active { background: #1a1a2e; color: #fff; border-color: #1a1a2e; }
"""

_JS = """
function sortTable(tableId, colIdx) {
    const table = document.getElementById(tableId);
    const tbody = table.querySelector('tbody');
    const rows = Array.from(tbody.querySelectorAll('tr.data-row'));
    const detailRows = {};
    rows.forEach(r => {
        const next = r.nextElementSibling;
        if (next && next.classList.contains('detail-row')) detailRows[r.dataset.task] = next;
    });
    const dir = table.dataset.sortCol == colIdx && table.dataset.sortDir == 'asc' ? 'desc' : 'asc';
    table.dataset.sortCol = colIdx;
    table.dataset.sortDir = dir;
    rows.sort((a, b) => {
        let va = a.children[colIdx]?.dataset.val ?? a.children[colIdx]?.textContent ?? '';
        let vb = b.children[colIdx]?.dataset.val ?? b.children[colIdx]?.textContent ?? '';
        let na = parseFloat(va), nb = parseFloat(vb);
        if (!isNaN(na) && !isNaN(nb)) return dir === 'asc' ? na - nb : nb - na;
        return dir === 'asc' ? va.localeCompare(vb) : vb.localeCompare(va);
    });
    rows.forEach(r => {
        tbody.appendChild(r);
        if (detailRows[r.dataset.task]) tbody.appendChild(detailRows[r.dataset.task]);
    });
}
function toggleDetail(task) {
    const row = document.querySelector(`tr.detail-row[data-task="${task}"]`);
    const icon = document.querySelector(`tr.data-row[data-task="${task}"] .expand-icon`);
    if (row) { row.classList.toggle('open'); }
    if (icon) { icon.classList.toggle('open'); }
}
function filterRows(tableId, filterType) {
    document.querySelectorAll('.filter-bar button').forEach(b => b.classList.remove('active'));
    event.target.classList.add('active');
    const tbody = document.getElementById(tableId).querySelector('tbody');
    tbody.querySelectorAll('tr.data-row').forEach(r => {
        const show = filterType === 'all'
            || (filterType === 'pass' && r.dataset.passed === 'true')
            || (filterType === 'fail' && r.dataset.passed === 'false')
            || (filterType === 'error' && r.dataset.status === 'error');
        r.style.display = show ? '' : 'none';
        const detail = r.nextElementSibling;
        if (detail && detail.classList.contains('detail-row')) {
            detail.style.display = show && detail.classList.contains('open') ? '' : 'none';
        }
    });
}
"""


def _esc(text: str) -> str:
    return html_mod.escape(str(text)) if text else ""


def _fmt_score(score) -> str:
    if score is None:
        return "-"
    return f"{float(score):.1f}"


def _fmt_cost(cost) -> str:
    if not cost:
        return "$0.00"
    return f"${float(cost):.2f}"


def _fmt_time(seconds) -> str:
    if not seconds:
        return "-"
    s = float(seconds)
    if s < 60:
        return f"{s:.0f}s"
    return f"{s / 60:.1f}m"


def _fmt_tokens(count) -> str:
    if not count:
        return "0"
    n = int(count)
    if n >= 1_000_000:
        return f"{n / 1_000_000:.1f}M"
    if n >= 1_000:
        return f"{n / 1_000:.1f}k"
    return str(n)


def _delta_html(val) -> str:
    if val is None:
        return "-"
    v = float(val)
    cls = "delta-pos" if v > 0 else "delta-neg" if v < 0 else ""
    sign = "+" if v > 0 else ""
    return f'<span class="{cls}">{sign}{v:.1f}</span>'


def generate_single_html(experiment: Dict[str, Any]) -> str:
    """Generate HTML dashboard for a single experiment."""
    meta = experiment["meta"]
    tasks = experiment["tasks"]
    exp_name = meta.get("experiment_name", "Experiment")
    config = meta.get("config", {})
    controller = config.get("controller", {})
    model = controller.get("args", {}).get("model", controller.get("type", "unknown"))
    agent_type = config.get("agent_type", "unknown")
    timestamp = meta.get("timestamp", "")
    summary = meta.get("summary", {})

    all_metrics = {name: extract_task_metrics(r) for name, r in tasks.items()}
    total_cost = sum(m["cost"] for m in all_metrics.values())
    total_time = sum(m["execution_time"] for m in all_metrics.values())
    total_input_tokens = sum(m["input_tokens"] for m in all_metrics.values())
    total_output_tokens = sum(m["output_tokens"] for m in all_metrics.values())
    total_tokens = sum(m["total_tokens"] for m in all_metrics.values())

    # Build table rows
    table_rows = ""
    for task_name in sorted(all_metrics.keys()):
        m = all_metrics[task_name]
        status_cls = "pass" if m["passed"] else ("error" if m["status"] == "error" else "fail")
        status_txt = "PASS" if m["passed"] else ("ERROR" if m["status"] == "error" else "FAIL")
        dims = m["dimensions"]
        dim_cells = "".join(
            f'<td data-val="{dims.get(d, "")}">{_fmt_score(dims.get(d))}</td>'
            for d in ["relevance", "accuracy", "cost_effectiveness", "justification"]
        )
        table_rows += f"""
        <tr class="data-row expandable" data-task="{_esc(task_name)}"
            data-passed="{str(m['passed']).lower()}" data-status="{m['status']}"
            onclick="toggleDetail('{_esc(task_name)}')">
            <td><span class="expand-icon">&#9654;</span> {_esc(task_name)}</td>
            <td class="{status_cls}">{status_txt}</td>
            <td data-val="{m['score'] if m['score'] is not None else ''}">{_fmt_score(m['score'])}</td>
            {dim_cells}
            <td data-val="{m['iterations']}">{m['iterations']}</td>
            <td data-val="{m['input_tokens']}">{_fmt_tokens(m['input_tokens'])}</td>
            <td data-val="{m['output_tokens']}">{_fmt_tokens(m['output_tokens'])}</td>
            <td data-val="{m['total_tokens']}">{_fmt_tokens(m['total_tokens'])}</td>
            <td data-val="{m['cost']}">{_fmt_cost(m['cost'])}</td>
            <td data-val="{m['execution_time']}">{_fmt_time(m['execution_time'])}</td>
        </tr>
        <tr class="detail-row" data-task="{_esc(task_name)}">
            <td colspan="13" class="detail-cell">
                <h4>Final Answer</h4>
                <pre>{_esc(m['task_result'][:2000])}</pre>
                <h4>Execution Summary (sandbox actions)</h4>
                <pre>{_esc(m['execution_summary'][:3000]) or '(no sandbox actions — agent answered directly)'}</pre>
                <h4>Full Conversation Trace</h4>
                <pre>{_esc(m['conversation_trace'][:8000])}</pre>
                <h4>Judge Reasoning</h4>
                <pre>{_esc(m['judge_reasoning'])}</pre>
            </td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>{_esc(exp_name)} — Experiment Report</title>
<style>{_CSS}</style></head><body>
<div class="header">
    <h1>{_esc(exp_name)}</h1>
    <p>Model: <strong>{_esc(model)}</strong> &nbsp;|&nbsp; Agent: <strong>{_esc(agent_type)}</strong>
       &nbsp;|&nbsp; {_esc(timestamp)}</p>
</div>
<div class="cards">
    <div class="card"><div class="label">Tasks</div><div class="value">{summary.get('passed', 0) + summary.get('failed', 0) + summary.get('errors', 0)}</div></div>
    <div class="card"><div class="label">Pass Rate</div><div class="value">{summary.get('pass_rate', 0):.1f}%</div></div>
    <div class="card"><div class="label">Avg Score</div><div class="value">{summary.get('avg_score', 0):.2f}</div></div>
    <div class="card"><div class="label">Total Tokens</div><div class="value">{_fmt_tokens(total_tokens)}</div></div>
    <div class="card"><div class="label">Input / Output</div><div class="value">{_fmt_tokens(total_input_tokens)} / {_fmt_tokens(total_output_tokens)}</div></div>
    <div class="card"><div class="label">Total Cost</div><div class="value">{_fmt_cost(total_cost)}</div></div>
    <div class="card"><div class="label">Total Time</div><div class="value">{_fmt_time(total_time)}</div></div>
</div>
<div class="filter-bar">
    <button class="active" onclick="filterRows('results','all')">All</button>
    <button onclick="filterRows('results','pass')">Passed</button>
    <button onclick="filterRows('results','fail')">Failed</button>
    <button onclick="filterRows('results','error')">Errors</button>
</div>
<table id="results">
<thead><tr>
    <th onclick="sortTable('results',0)">Task</th>
    <th onclick="sortTable('results',1)">Status</th>
    <th onclick="sortTable('results',2)">Score</th>
    <th onclick="sortTable('results',3)">Relevance</th>
    <th onclick="sortTable('results',4)">Accuracy</th>
    <th onclick="sortTable('results',5)">Cost Eff.</th>
    <th onclick="sortTable('results',6)">Justification</th>
    <th onclick="sortTable('results',7)">Iterations</th>
    <th onclick="sortTable('results',8)">In Tokens</th>
    <th onclick="sortTable('results',9)">Out Tokens</th>
    <th onclick="sortTable('results',10)">Total Tokens</th>
    <th onclick="sortTable('results',11)">Cost</th>
    <th onclick="sortTable('results',12)">Time</th>
</tr></thead>
<tbody>{table_rows}</tbody>
</table>
<script>{_JS}</script>
</body></html>"""


def generate_comparison_html(experiments: List[Dict[str, Any]]) -> str:
    """Generate HTML dashboard comparing N experiments."""
    exp_names = [e["meta"].get("experiment_name", f"Exp {i+1}") for i, e in enumerate(experiments)]

    all_task_names = sorted(set(
        tn for e in experiments for tn in e["tasks"].keys()
    ))

    # Per-experiment summaries
    exp_summaries = []
    for exp in experiments:
        metrics = {tn: extract_task_metrics(r) for tn, r in exp["tasks"].items()}
        scores = [m["score"] for m in metrics.values() if m["score"] is not None]
        passed = sum(1 for m in metrics.values() if m["passed"])
        total = len(metrics)
        total_cost = sum(m["cost"] for m in metrics.values())
        total_time = sum(m["execution_time"] for m in metrics.values())
        total_in_tok = sum(m["input_tokens"] for m in metrics.values())
        total_out_tok = sum(m["output_tokens"] for m in metrics.values())
        total_tok = sum(m["total_tokens"] for m in metrics.values())
        exp_summaries.append({
            "name": exp["meta"].get("experiment_name", "?"),
            "model": exp["meta"].get("config", {}).get("controller", {}).get("args", {}).get("model", "?"),
            "agent_type": exp["meta"].get("config", {}).get("agent_type", "?"),
            "total": total,
            "passed": passed,
            "pass_rate": 100 * passed / total if total else 0,
            "avg_score": sum(scores) / len(scores) if scores else 0,
            "total_cost": total_cost,
            "total_time": total_time,
            "total_input_tokens": total_in_tok,
            "total_output_tokens": total_out_tok,
            "total_tokens": total_tok,
        })

    # Summary cards
    summary_cards = ""
    for s in exp_summaries:
        summary_cards += f"""
        <div class="card">
            <div class="label">{_esc(s['name'])}</div>
            <div class="value">{s['pass_rate']:.1f}%</div>
            <p style="font-size:.8rem;color:#666;margin-top:4px">
                Score: {s['avg_score']:.2f} &nbsp;|&nbsp; Cost: {_fmt_cost(s['total_cost'])}
                &nbsp;|&nbsp; Tokens: {_fmt_tokens(s['total_tokens'])} ({_fmt_tokens(s['total_input_tokens'])} in / {_fmt_tokens(s['total_output_tokens'])} out)
                &nbsp;|&nbsp; {_esc(s['model'])} / {_esc(s['agent_type'])}
            </p>
        </div>"""

    # Build column headers
    n_exp = len(experiments)
    header_cols = '<th onclick="sortTable(\'results\',0)">Task</th>'
    col_idx = 1
    for i, name in enumerate(exp_names):
        header_cols += f'<th onclick="sortTable(\'results\',{col_idx})">{_esc(name)}<br>Status</th>'
        col_idx += 1
        header_cols += f'<th onclick="sortTable(\'results\',{col_idx})">{_esc(name)}<br>Score</th>'
        col_idx += 1
        header_cols += f'<th onclick="sortTable(\'results\',{col_idx})">{_esc(name)}<br>Iters</th>'
        col_idx += 1
        header_cols += f'<th onclick="sortTable(\'results\',{col_idx})">{_esc(name)}<br>Tokens</th>'
        col_idx += 1
        header_cols += f'<th onclick="sortTable(\'results\',{col_idx})">{_esc(name)}<br>Cost</th>'
        col_idx += 1
        if i > 0:
            header_cols += f'<th onclick="sortTable(\'results\',{col_idx})">Score &Delta;</th>'
            col_idx += 1
            header_cols += f'<th onclick="sortTable(\'results\',{col_idx})">Token &Delta;</th>'
            col_idx += 1

    # Build table rows
    table_rows = ""
    for task_name in all_task_names:
        metrics_per_exp = []
        for exp in experiments:
            if task_name in exp["tasks"]:
                metrics_per_exp.append(extract_task_metrics(exp["tasks"][task_name]))
            else:
                metrics_per_exp.append(None)

        first_valid = next((m for m in metrics_per_exp if m), None)
        data_passed = str(first_valid["passed"]).lower() if first_valid else "false"
        data_status = first_valid["status"] if first_valid else "unknown"

        cells = f'<td><span class="expand-icon">&#9654;</span> {_esc(task_name)}</td>'
        prev_score = None
        prev_tokens = None
        for i, m in enumerate(metrics_per_exp):
            if m is None:
                cells += '<td>-</td><td>-</td><td>-</td><td>-</td><td>-</td>'
                if i > 0:
                    cells += '<td>-</td><td>-</td>'
                continue
            status_cls = "pass" if m["passed"] else ("error" if m["status"] == "error" else "fail")
            status_txt = "PASS" if m["passed"] else ("ERROR" if m["status"] == "error" else "FAIL")
            cells += f'<td class="{status_cls}">{status_txt}</td>'
            cells += f'<td data-val="{m["score"] if m["score"] is not None else ""}">{_fmt_score(m["score"])}</td>'
            cells += f'<td data-val="{m["iterations"]}">{m["iterations"]}</td>'
            cells += f'<td data-val="{m["total_tokens"]}" title="In: {_fmt_tokens(m["input_tokens"])} / Out: {_fmt_tokens(m["output_tokens"])}">{_fmt_tokens(m["total_tokens"])}</td>'
            cells += f'<td data-val="{m["cost"]}">{_fmt_cost(m["cost"])}</td>'
            if i > 0:
                score_delta = None
                if m["score"] is not None and prev_score is not None:
                    score_delta = float(m["score"]) - float(prev_score)
                cells += f'<td>{_delta_html(score_delta)}</td>'
                token_delta = None
                if m["total_tokens"] and prev_tokens:
                    token_delta = m["total_tokens"] - prev_tokens
                token_delta_str = "-"
                if token_delta is not None:
                    sign = "+" if token_delta > 0 else ""
                    cls = "delta-neg" if token_delta > 0 else ("delta-pos" if token_delta < 0 else "")
                    token_delta_str = f'<span class="{cls}">{sign}{_fmt_tokens(token_delta)}</span>'
                cells += f'<td>{token_delta_str}</td>'
            prev_score = m["score"]
            prev_tokens = m["total_tokens"]

        # Detail row with side-by-side info
        detail_cols = n_exp
        grid_cols = f"grid-template-columns: repeat({detail_cols}, 1fr)"
        detail_panels = ""
        for i, (ename, m) in enumerate(zip(exp_names, metrics_per_exp)):
            if m is None:
                detail_panels += f'<div><div class="exp-label">{_esc(ename)}</div><p>No result</p></div>'
                continue
            skills_text = ""
            if m["skills_used"]:
                skills_text = f'<h4>Skills Used</h4><pre>{_esc(", ".join(m["skills_used"]))}</pre>'
            detail_panels += f"""
            <div>
                <div class="exp-label">{_esc(ename)}</div>
                <h4>Final Answer</h4>
                <pre>{_esc((m['task_result'] or '')[:1500])}</pre>
                <h4>Execution Summary (sandbox actions)</h4>
                <pre>{_esc((m['execution_summary'] or '')[:2000]) or '(no sandbox actions)'}</pre>
                <h4>Full Conversation Trace</h4>
                <pre>{_esc((m.get('conversation_trace') or '')[:5000])}</pre>
                <h4>Judge Reasoning</h4>
                <pre>{_esc(m['judge_reasoning'] or '')}</pre>
                {skills_text}
            </div>"""

        total_cols = 1 + n_exp * 5 + (n_exp - 1) * 2
        table_rows += f"""
        <tr class="data-row expandable" data-task="{_esc(task_name)}"
            data-passed="{data_passed}" data-status="{data_status}"
            onclick="toggleDetail('{_esc(task_name)}')">
            {cells}
        </tr>
        <tr class="detail-row" data-task="{_esc(task_name)}">
            <td colspan="{total_cols}" class="detail-cell">
                <div class="side-by-side" style="{grid_cols}">
                    {detail_panels}
                </div>
            </td>
        </tr>"""

    header_info = " vs ".join(f"<strong>{_esc(n)}</strong>" for n in exp_names)

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Experiment Comparison</title>
<style>{_CSS}</style></head><body>
<div class="header">
    <h1>Experiment Comparison</h1>
    <p>{header_info}</p>
</div>
<div class="cards">{summary_cards}</div>
<div class="filter-bar">
    <button class="active" onclick="filterRows('results','all')">All</button>
    <button onclick="filterRows('results','pass')">Passed</button>
    <button onclick="filterRows('results','fail')">Failed</button>
    <button onclick="filterRows('results','error')">Errors</button>
</div>
<table id="results">
<thead><tr>{header_cols}</tr></thead>
<tbody>{table_rows}</tbody>
</table>
<script>{_JS}</script>
</body></html>"""


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def generate_single_excel(experiment: Dict[str, Any], output_path: Path):
    """Generate Excel workbook for a single experiment."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    header_font = Font(bold=True)
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    meta = experiment["meta"]
    summary = meta.get("summary", {})
    info = [
        ("Experiment", meta.get("experiment_name", "")),
        ("Model", meta.get("config", {}).get("controller", {}).get("args", {}).get("model", "")),
        ("Agent Type", meta.get("config", {}).get("agent_type", "")),
        ("Tasks Dir", meta.get("tasks_dir", "")),
        ("Timestamp", meta.get("timestamp", "")),
        ("Total Tasks", summary.get("passed", 0) + summary.get("failed", 0) + summary.get("errors", 0)),
        ("Passed", summary.get("passed", 0)),
        ("Failed", summary.get("failed", 0)),
        ("Errors", summary.get("errors", 0)),
        ("Pass Rate", f"{summary.get('pass_rate', 0)}%"),
        ("Avg Score", summary.get("avg_score", 0)),
    ]
    for row_idx, (label, val) in enumerate(info, 1):
        ws.cell(row=row_idx, column=1, value=label).font = header_font
        ws.cell(row=row_idx, column=2, value=val)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40

    # --- Per-Task sheet ---
    ws2 = wb.create_sheet("Per-Task")
    headers = ["Task", "Status", "Passed", "Score", "Relevance", "Accuracy",
               "Cost Eff.", "Justification", "Iterations",
               "Input Tokens", "Output Tokens", "Cached Tokens", "Total Tokens", "API Calls",
               "Cost ($)", "Time (s)"]
    for ci, h in enumerate(headers, 1):
        ws2.cell(row=1, column=ci, value=h).font = header_font

    for ri, (task_name, result) in enumerate(sorted(experiment["tasks"].items()), 2):
        m = extract_task_metrics(result)
        dims = m["dimensions"]
        row = [task_name, m["status"], m["passed"], m["score"],
               dims.get("relevance"), dims.get("accuracy"),
               dims.get("cost_effectiveness"), dims.get("justification"),
               m["iterations"],
               m["input_tokens"], m["output_tokens"], m["cached_tokens"], m["total_tokens"], m["api_calls"],
               round(m["cost"], 4), round(m["execution_time"], 1)]
        for ci, val in enumerate(row, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            if ci == 3:
                cell.fill = pass_fill if val else fail_fill

    for ci in range(1, len(headers) + 1):
        ws2.column_dimensions[chr(64 + ci)].width = 14
    ws2.column_dimensions["A"].width = 30

    # --- Traces sheet ---
    ws3 = wb.create_sheet("Traces")
    ws3.cell(row=1, column=1, value="Task").font = header_font
    ws3.cell(row=1, column=2, value="Execution Summary").font = header_font
    ws3.cell(row=1, column=3, value="Final Answer").font = header_font
    ws3.cell(row=1, column=4, value="Judge Reasoning").font = header_font

    for ri, (task_name, result) in enumerate(sorted(experiment["tasks"].items()), 2):
        m = extract_task_metrics(result)
        ws3.cell(row=ri, column=1, value=task_name)
        ws3.cell(row=ri, column=2, value=(m["execution_summary"] or "")[:32000])
        ws3.cell(row=ri, column=3, value=(m["task_result"] or "")[:32000])
        ws3.cell(row=ri, column=4, value=m["judge_reasoning"] or "")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 80
    ws3.column_dimensions["C"].width = 60
    ws3.column_dimensions["D"].width = 60

    wb.save(output_path)


def generate_comparison_excel(experiments: List[Dict[str, Any]], output_path: Path):
    """Generate Excel workbook comparing N experiments."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    header_font = Font(bold=True)
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    pos_fill = PatternFill("solid", fgColor="C6EFCE")
    neg_fill = PatternFill("solid", fgColor="FFC7CE")

    exp_names = [e["meta"].get("experiment_name", f"Exp {i+1}") for i, e in enumerate(experiments)]

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    sum_headers = ["Experiment", "Model", "Agent", "Tasks", "Passed", "Failed",
                   "Errors", "Pass Rate", "Avg Score",
                   "Input Tokens", "Output Tokens", "Total Tokens",
                   "Total Cost ($)", "Total Time (s)"]
    for ci, h in enumerate(sum_headers, 1):
        ws.cell(row=1, column=ci, value=h).font = header_font

    for ri, exp in enumerate(experiments, 2):
        meta = exp["meta"]
        cfg = meta.get("config", {})
        metrics = {tn: extract_task_metrics(r) for tn, r in exp["tasks"].items()}
        scores = [m["score"] for m in metrics.values() if m["score"] is not None]
        passed = sum(1 for m in metrics.values() if m["passed"])
        total = len(metrics)
        total_cost = sum(m["cost"] for m in metrics.values())
        total_time = sum(m["execution_time"] for m in metrics.values())
        total_in_tok = sum(m["input_tokens"] for m in metrics.values())
        total_out_tok = sum(m["output_tokens"] for m in metrics.values())
        total_tok = sum(m["total_tokens"] for m in metrics.values())
        row = [
            meta.get("experiment_name", ""),
            cfg.get("controller", {}).get("args", {}).get("model", ""),
            cfg.get("agent_type", ""),
            total, passed, total - passed - sum(1 for m in metrics.values() if m["status"] == "error"),
            sum(1 for m in metrics.values() if m["status"] == "error"),
            f"{100 * passed / total:.1f}%" if total else "0%",
            round(sum(scores) / len(scores), 2) if scores else 0,
            total_in_tok, total_out_tok, total_tok,
            round(total_cost, 2),
            round(total_time, 1),
        ]
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    for ci in range(1, len(sum_headers) + 1):
        ws.column_dimensions[chr(64 + min(ci, 26))].width = 16

    # --- Per-Task sheet ---
    ws2 = wb.create_sheet("Per-Task")
    pt_headers = ["Task"]
    for i, name in enumerate(exp_names):
        pt_headers += [f"{name} Status", f"{name} Score", f"{name} Iters",
                       f"{name} In Tokens", f"{name} Out Tokens", f"{name} Total Tokens",
                       f"{name} Cost ($)"]
        if i > 0:
            pt_headers.append(f"Score Delta ({exp_names[i-1]} → {name})")
            pt_headers.append(f"Token Delta ({exp_names[i-1]} → {name})")
    for ci, h in enumerate(pt_headers, 1):
        ws2.cell(row=1, column=ci, value=h).font = header_font

    all_task_names = sorted(set(tn for e in experiments for tn in e["tasks"].keys()))
    for ri, task_name in enumerate(all_task_names, 2):
        ws2.cell(row=ri, column=1, value=task_name)
        ci = 2
        prev_score = None
        prev_tokens = None
        for ei, exp in enumerate(experiments):
            if task_name in exp["tasks"]:
                m = extract_task_metrics(exp["tasks"][task_name])
                ws2.cell(row=ri, column=ci, value="PASS" if m["passed"] else "FAIL")
                ws2.cell(row=ri, column=ci).fill = pass_fill if m["passed"] else fail_fill
                ci += 1
                ws2.cell(row=ri, column=ci, value=m["score"])
                ci += 1
                ws2.cell(row=ri, column=ci, value=m["iterations"])
                ci += 1
                ws2.cell(row=ri, column=ci, value=m["input_tokens"])
                ci += 1
                ws2.cell(row=ri, column=ci, value=m["output_tokens"])
                ci += 1
                ws2.cell(row=ri, column=ci, value=m["total_tokens"])
                ci += 1
                ws2.cell(row=ri, column=ci, value=round(m["cost"], 4))
                ci += 1
                if ei > 0:
                    score_delta = None
                    if m["score"] is not None and prev_score is not None:
                        score_delta = round(float(m["score"]) - float(prev_score), 1)
                    cell = ws2.cell(row=ri, column=ci, value=score_delta)
                    if score_delta is not None:
                        cell.fill = pos_fill if score_delta > 0 else (neg_fill if score_delta < 0 else PatternFill())
                    ci += 1
                    token_delta = None
                    if m["total_tokens"] and prev_tokens:
                        token_delta = m["total_tokens"] - prev_tokens
                    cell = ws2.cell(row=ri, column=ci, value=token_delta)
                    if token_delta is not None:
                        cell.fill = pos_fill if token_delta < 0 else (neg_fill if token_delta > 0 else PatternFill())
                    ci += 1
                prev_score = m["score"]
                prev_tokens = m["total_tokens"]
            else:
                ws2.cell(row=ri, column=ci, value="-"); ci += 1
                ws2.cell(row=ri, column=ci, value="-"); ci += 1
                ws2.cell(row=ri, column=ci, value="-"); ci += 1
                ws2.cell(row=ri, column=ci, value="-"); ci += 1
                ws2.cell(row=ri, column=ci, value="-"); ci += 1
                ws2.cell(row=ri, column=ci, value="-"); ci += 1
                ws2.cell(row=ri, column=ci, value="-"); ci += 1
                if ei > 0:
                    ws2.cell(row=ri, column=ci, value="-"); ci += 1
                    ws2.cell(row=ri, column=ci, value="-"); ci += 1
                prev_score = None
                prev_tokens = None

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Generate HTML + Excel reports from experiment result directories."
    )
    parser.add_argument("dirs", nargs="+", type=str,
                        help="One or more experiment result directories")
    parser.add_argument("--output", "-o", type=str, default="",
                        help="Output directory (defaults to experiment dir for single, CWD for comparison)")
    args = parser.parse_args()

    dirs = [Path(d) for d in args.dirs]
    for d in dirs:
        if not d.is_dir():
            print(f"[error] Not a directory: {d}")
            sys.exit(1)

    experiments = [load_experiment(d) for d in dirs]

    if len(experiments) == 1:
        exp = experiments[0]
        out_dir = Path(args.output) if args.output else dirs[0]
        out_dir.mkdir(parents=True, exist_ok=True)

        html_path = out_dir / "report.html"
        with open(html_path, "w") as f:
            f.write(generate_single_html(exp))
        print(f"HTML report: {html_path}")

        try:
            xlsx_path = out_dir / "report.xlsx"
            generate_single_excel(exp, xlsx_path)
            print(f"Excel report: {xlsx_path}")
        except ImportError:
            print("[warn] openpyxl not installed, skipping Excel. Install with: pip install openpyxl")
    else:
        out_dir = Path(args.output) if args.output else Path(".")
        out_dir.mkdir(parents=True, exist_ok=True)

        html_path = out_dir / "comparison_report.html"
        with open(html_path, "w") as f:
            f.write(generate_comparison_html(experiments))
        print(f"HTML comparison: {html_path}")

        try:
            xlsx_path = out_dir / "comparison_report.xlsx"
            generate_comparison_excel(experiments, xlsx_path)
            print(f"Excel comparison: {xlsx_path}")
        except ImportError:
            print("[warn] openpyxl not installed, skipping Excel. Install with: pip install openpyxl")


if __name__ == "__main__":
    main()
